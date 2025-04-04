import csv
import os
from matplotlib import pyplot as plt
from matplotlib.ticker import MaxNLocator
import mplcursors
import copy
import re
import json
import pandas as pd
import numpy as np
import seaborn as sns
import operator
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

###########
# Globals #
###########

__version__ = "0.0.6"

# Labels for different types of molecular interactions
INTERACTION_LABELS = [
    "Hydrophobic", "Aromatic_Face/Face", "Aromatic_Edge/Face", "HBond_PROT", "HBond_LIG", 
    "Ionic_PROT", "Ionic_LIG", "Metal Acceptor", "Pi/Cation", "Other_Interactions"
]

# List of colors corresponding to different interaction types
COLORS = [
    "#ff6384", "#36a2eb", "#ffce56", "#4bc0c0", "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728",
    "#9467bd", "#8c564b"
]

# List of available program modes
PROGRAM_MODES = [
    "ichem", "arpeggio"
]

# Available modes for heatmap visualization
HEATMAP_MODES = [
    "max", "min", "mean", "count", "percent"
]

# Arpeggio-specific interaction entities
ARPEGGIO_INT_ENT = [
    "INTER"
]

# Types of molecular contacts identified by Arpeggio
ARPEGGIO_CONT = [
    "covalent", "hbond", "aromatic", "hydrophobic", "polar", "ionic", "xbond", "metal", 
    "carbonyl", "CARBONPI", "CATIONPI", "DONORPI", "HALOGENPI", "METSULPHURPI", 
    "AMIDEAMIDE", "AMIDERING"
]

# Types of Arpeggio interactions (currently only plane-plane interactions)
ARPEGGIO_TYPE = [
    "plane-plane"
]

# Colors associated with Arpeggio interaction types
ARPEGGIO_COLORS = [
    "#e6194B", "#3cb44b", "#ffe119", "#4363d8", "#f58231", "#911eb4",
    "#46f0f0", "#f032e6", "#bcf60c", "#fabebe", "#008080", "#e6beff",
    "#9a6324", "#fffac8", "#800000", "#aaffc3", "#808000"
]

# Constants for delimiters used in interaction data representation
SAME_DELIM = ', '       # Separates interactions of the same type.
DIFF_DELIM = '; '       # Separates interactions of different types.
GROUP_DELIM = '|'       # Groups interactions of the same type.

# Constants for handling empty cell values in interaction matrices
EMPTY_CELL = ''         # Represents an originally empty cell.
EMPTY_DASH_CELL = '-'   # Represents a cell that is considered empty after filtering.

# Global list of three-letter amino acid codes
AMINO_ACID_CODES = [
    "ALA", "ARG", "ASN", "ASP", "CYS", "GLN", "GLU", "GLY", "HIS", "ILE", "LEU", "LYS", "MET", 
    "PHE", "PRO", "SER", "THR", "TRP", "TYR", "VAL"
]

###########
# Lambdas #
###########

is_not_empty_or_dash = lambda cell: not (cell == EMPTY_DASH_CELL or cell == EMPTY_CELL)

class InteractionData:
    def __init__(self, colors, interactions, ligand, matrix, mode, protein, subunit, subunits_set):
        self.colors = colors
        self.interactions = interactions
        self.ligand = ligand
        self.matrix = matrix
        self.mode = mode
        self.protein = protein
        self.subunit = subunit
        self.subunits_set = subunits_set

    def compare(self, other):
        # En lugar de isinstance, verificamos que tenga los mismos atributos
        if not hasattr(other, "__dict__"):
            return "Invalid comparison. The object does not has the espected attributes."

        differences = {}
        for attr in vars(self):  # Obtener todos los atributos del objeto
            if getattr(self, attr) != getattr(other, attr):
                differences[attr] = (getattr(self, attr), getattr(other, attr))

        return differences if differences else "There are no changes."

class AnalyzeInteractions:

    ROWS = "rows"
    COLUMNS = "columns"

    ARPEGGIO = "arpeggio"
    ICHEM = "ichem"

    INPUT = "input"
    OUTPUT = "output"

    MINIMUM = "min"
    MAXIMUM = "max"
    MEAN = "mean"
    COUNT = "count"
    PERCENT = "percent"

    MAIN = "<main>"
    SIDE = "<side>"

    LOWER = "lower"
    UPPER = "upper"

    def __init__(self):
        """Initializes the class with the current working directory and default settings."""
        self.saving_directory = os.getcwd()  # Set the default saving directory
        self.input_directory = os.getcwd()  # Set the default input directory
        self.interaction_labels = INTERACTION_LABELS  # Default interaction labels
        self.codes = True
        self.plot_colors = COLORS  # Default color configuration
        self.plot_max_cols = 80
        self.aa = AMINO_ACID_CODES
        self.heat_max_cols = 30
        self.heat_colors = "RdYlGn"

    ###################
    # Private Methods #
    ###################

    def _check_variable_types(self, variables: list, expected_types: list, variable_names: list[str]):
        """
        Check if variables match their expected types.

        Args:
            variables (list): List of variables to check.
            expected_types (list): List of expected types (can include tuples).
            variable_names (list[str]): List of variable names for error messages.

        Raises:
            ValueError: If the lengths of input lists don't match.
            TypeMismatchException: If a variable type doesn't match the expected one.
        """
        if len(variables) != len(expected_types) or len(variables) != len(variable_names):
            raise ValueError("The lists of variables, expected types, and variable names must all have the same length.")

        for i, variable in enumerate(variables):
            expected_type = expected_types[i]
            variable_name = variable_names[i]

            if not isinstance(variable, expected_type if isinstance(expected_type, tuple) else (expected_type,)):
                actual_type = type(variable)
                raise TypeMismatchException(variable_name, expected_type if isinstance(expected_type, tuple) else (expected_type,), actual_type)

    def _get_residues_axis(self, matrix: list[list[str]]) -> str:
        """
        Determine whether residues' axis is in rows or columns.

        Args:
            matrix (list[list[str]]): Matrix with interaction data.

        Returns:
            str: 'columns' if residues are in columns, 'rows' otherwise.

        Raises:
            ValueError: If the axis cannot be determined.
        """
        self._verify_dimensions(matrix=matrix)

        # Check '(' occurrence in specific cells to determine axis
        count_0_1 = matrix[0][1].count('(')
        count_1_0 = matrix[1][0].count('(')

        if count_0_1 == count_1_0: # This checks in case activity is used or not
            count_0_1 = matrix[0][1].count(', ')
            count_1_0 = matrix[1][0].count(', ')
            if count_0_1 == count_1_0: # This checks in case PDB and ligand code are used or not
                count_0_1 = matrix[0][1].count(' ')
                count_1_0 = matrix[1][0].count(' ')
                if count_0_1 == count_1_0:
                    raise ValueError("Cannot determine the residues' axis.")
                elif count_0_1 == 1:
                    return 'columns'
                elif count_1_0 == 1:
                    return 'rows'
                else:
                    raise ValueError("Cannot determine the residues' axis.")
            elif count_0_1 == 1:
                return 'rows'
            elif count_1_0 == 1:
                return 'columns'
            else:
                raise ValueError("Cannot determine the residues' axis.")
        elif count_0_1 == 1:
            return 'rows'
        elif count_1_0 == 1:
            return 'columns'
        else:
            raise ValueError("Cannot determine the residues' axis.")

    def _verify_dimensions(self, matrix: list[list[str]]) -> None:
        """
        Verify matrix dimensions to ensure at least 2 rows and 2 columns.

        Args:
            matrix (list[list[str]]): Matrix to verify.
        
        Returns:
            None

        Raises:
            ValueError: If the matrix is too small or any row is too short.
        """
        if len(matrix) < 2 or any(len(row) < 2 for row in matrix):
            raise ValueError("The matrix must have at least 2 rows and 2 columns.")

    def _get_interactions(self, cell: str) -> list[int]:
        """
        Extracts and counts interactions from a cell string.

        Args:
            cell (str): The cell string containing interaction data.

        Returns:
            list: A list of interaction counts for each interaction type.
        """
        interactions = [0] * len(self.interaction_labels)
        sections = cell.split(GROUP_DELIM)
        for index in range(1, len(sections), 2):
            interaction = int(sections[index - 1].replace(DIFF_DELIM, "").replace(" ", ""))
            interactions[interaction - 1] += len(sections[index].split(SAME_DELIM))
        return interactions

    def _stack_reactives(self, 
                         matrix: list[list[str]], 
                         axis: str, 
                         type_count: bool) -> tuple[list[list[int]], list[str]]:
        """
        Accumulates interaction counts for rows or columns and returns the stacked data.

        Args:
            matrix (list of lists): The matrix containing interaction data.
            axis (str): Specifies whether to select rows ('rows') or columns ('columns').
            type_count (bool): Whether to count types or instances.

        Returns:
            tuple: A tuple containing the stacked data and indices.
        """
        self._verify_dimensions(matrix=matrix)
        if axis == 'columns':
            matrix = self.transpose_matrix(matrix)

        reactives = {row: [0] * len(self.interaction_labels) for row in range(1, len(matrix))}
        indices = [matrix[row][0].split('_')[0].strip() for row in range(1, len(matrix))]

        for row in range(1, len(matrix)):
            for column in range(1, len(matrix[row])):
                cell = matrix[row][column]
                interactions = self._get_interactions(cell)
                for i in range(len(self.interaction_labels)):
                    if type_count:
                        reactives[row][i] += interactions[i]
                    elif interactions[i] > 0:
                        reactives[row][i] += 1

        result_list = list(reactives.values())
        return result_list, indices

    def _plot_init(self, colors, matrix, axis, type_count):
        if colors is None:
            colors = self.plot_colors

        # Ensure the number of colors matches the number of interaction labels
        if len(colors) < len(self.interaction_labels):
            raise ValueError(f"Not enough colors provided. Expected at least {len(self.interaction_labels)} colors, but got {len(colors)}.")

        # Calculate stacked data if necessary
        data, indices = self._stack_reactives(matrix=matrix,
                                              axis=axis,
                                              type_count=type_count)
        transposed_data = self.transpose_matrix(data)
        return colors, data, indices, transposed_data

    def _plot_end(self, save, plt, fig, plot_name):
        # Show or save the plot
        if not save:
            plt.show()
        else:
            plt.savefig(os.path.join(self.saving_directory, plot_name + '.png'))
            plt.close(fig)  # Close the figure after saving to avoid display overlap

    def _verify_font(self, font: str) -> str:
        """
        Verifies the font parameter for heatmap visualization.

        Args:
            font (str): The font type to verify.

        Returns:
            str: The verified font type.

        Raises:
            ValueError: If the font is not recognized.
        """
        if font not in [self.LOWER, self.UPPER, None]:
            raise ValueError(f"Invalid font '{font}'. Expected 'lower', 'upper' or None.")
        return font
    
    #################################
    # Public Methods: Configuration #
    #################################

    def change_directory(
            self, 
            path: str,
            mode: str
            ) -> None:
        """
        Changes the working directory for saving or input operations.

        Args:
            path (str): Name of the subdirectory to switch to.
            mode (str): Determines whether to change the input or output directory.
                - 'input': Sets the directory for input files.
                - 'output': Sets the directory for output files.

        Returns:
            None

        Raises:
            ValueError: If the specified directory does not exist.
            InvalidModeException: If an invalid mode is provided.
        """
        
        self._check_variable_types(
            variables=[path, mode],
            expected_types=[str, str],
            variable_names=['path', 'mode']
        )

        # Construct the new path
        new_path = os.path.join(os.getcwd(), path)

        # Verify the new path exists
        if not os.path.exists(new_path):
            raise ValueError("The specified directory must exist inside the project.")

        if mode == 'input':
            self.input_directory = new_path
        elif mode == 'output':
            self.saving_directory = new_path
        else:
            raise InvalidModeException(mode=mode, expected_values=['input', 'output'])

    def set_config(
            self, 
            interactions: list[str] = None, 
            plot_max_cols: int = None,
            plot_colors: list[str] = None, 
            reset: bool = False,
            mode: str = None,
            heat_max_cols: int = None,
            heat_colors: str = None,
            interaction_data: InteractionData = None
            ) -> None:
        """
        Configures interaction settings, including labels, colors, and visualization parameters.

        Args:
            interactions (list[str], optional): List of interaction labels.
            plot_max_cols (int, optional): Maximum number of columns for plot visualization.
            plot_colors (list[str], optional): List of colors in hexadecimal format.
            reset (bool, optional): If True, resets configurations to default values.
            mode (str, optional): Determines preset configurations for different analysis modes ('ichem' or 'arpeggio').
            heat_max_cols (int, optional): Maximum number of columns for heatmap visualization.
            heat_colors (str, optional): Color scheme for heatmaps.
            interaction_data (InteractionData, optional): Object containing interaction settings to be applied.

        Returns:
            None

        Raises:
            InvalidColorException: If any color in the provided list is not a valid hexadecimal value.
            InvalidModeException: If an invalid mode is provided.
        """

        def is_valid_hex_color(color: str) -> bool:
            """
            Validates if a string is a valid hexadecimal color.

            Args:
                color (str): Color string to validate.

            Returns:
                bool: True if the color is a valid hex value, False otherwise.
            """
            return bool(re.match(r'^#([0-9a-fA-F]{6}|[0-9a-fA-F]{3})$', color))

        def reset_configuration() -> None:
            """
            Resets interaction labels and colors to their default values.

            Returns:
                None
            """
            self.saving_directory = os.getcwd()  # Set the default saving directory
            self.input_directory = os.getcwd()  # Set the default input directory
            self.interaction_labels = INTERACTION_LABELS  # Default interaction labels
            self.codes = True
            self.plot_colors = COLORS  # Default color configuration
            self.plot_max_cols = 80
            self.aa = AMINO_ACID_CODES
            self.heat_max_cols = 30
            self.heat_colors = "RdYlGn"

        self._check_variable_types(
            variables=[interactions, plot_max_cols, plot_colors, reset, mode, heat_max_cols, heat_colors, interaction_data],
            expected_types=[(list, None.__class__), (int, None.__class__), (list, None.__class__), bool, (str, None.__class__), (int, None.__class__), (str, None.__class__), (InteractionData, None.__class__)],
            variable_names=['interactions', 'plot_max_cols', 'colors', 'reset', 'mode', 'heat_max_elems', 'heat_colors', 'interaction_data']
        )

        # Perform reset if requested
        if reset:
            reset_configuration()
            return

        # Update interaction labels if provided
        if interactions:
            self.interaction_labels = interactions

        # Update colors if provided and valid
        if plot_colors:
            invalid_colors = [color for color in plot_colors if not is_valid_hex_color(color)]
            if invalid_colors:
                raise InvalidColorException(invalid_colors)
            self.plot_colors = plot_colors

        if mode:
            if mode == 'ichem':
                self.interaction_labels = INTERACTION_LABELS  # Default interaction labels
                self.plot_colors = COLORS  # Default color configuration
            elif mode == 'arpeggio':
                self.interaction_labels = ARPEGGIO_CONT + ARPEGGIO_TYPE  # Default interaction labels
                self.plot_colors = ARPEGGIO_COLORS  # Default color c
            else:
                raise InvalidModeException(mode=mode, expected_values=PROGRAM_MODES)
        if plot_max_cols:
            self.plot_max_cols = plot_max_cols
        if heat_max_cols:
            self.heat_max_cols = heat_max_cols
        if heat_colors:
            self.heat_colors = heat_colors
        if interaction_data:
            self.interaction_labels = interaction_data.interactions
            self.plot_colors = interaction_data.colors

    #################################
    # Public Methods: Functionality #
    #################################

    def analyze_files(self, 
        directory: str, 
        mode: str, 
        activity_file: str = None, 
        protein: bool = True, 
        ligand: bool = True, 
        subunit: bool = False, 
        save: str = None
    ) -> InteractionData:
        """
        Analyzes interaction data files in a specified directory, processing them according to the specified mode.

        This function reads interaction data files, extracts relevant interaction details, and organizes them into a
        structured matrix format. Optionally, an activity file can be provided to label data based on activity levels.

        Args:
            directory (str): Path to the directory containing interaction data files.
            mode (str): Processing mode. Supported modes:
                - 'ichem': Processes IChem interaction files.
                - 'arpeggio': Processes Arpeggio interaction files.
            activity_file (str, optional): Path to a CSV file containing activity data for annotation.
            protein (bool, optional): Whether to include protein atoms in the analysis. Defaults to True.
            ligand (bool, optional): Whether to include ligand atoms in the analysis. Defaults to True.
            subunit (bool, optional): Whether to differentiate between protein subunits. Defaults to False.
            save (str, optional): Path to save the processed interaction matrix. Defaults to None.

        Returns:
            InteractionData: An object containing the processed interaction matrix, metadata, and interaction labels.

        Raises:
            FileNotFoundError: If the specified directory or activity file does not exist.
            EmptyDirectoryException: If the specified directory contains no valid files.
            InvalidModeException: If the provided mode is not supported.
            InvalidFilenameException: If any file names contain spaces or invalid characters.
        """

        def label_matrix(
            matrix: list[list[str]], 
            rows: list[str], 
            columns: list[str], 
            activity_file: str,
            correction: list[str] = None
        ) -> list[list[str]]:
            """
            Adds appropriate headers to the interaction matrix, including residue and file names.

            If an activity file is provided, it also labels the columns with corresponding activity values.

            Args:
                matrix (list[list[str]]): 2D list representing interaction data.
                rows (list[str]): List of residue names for row labeling.
                columns (list[str]): List of file names for column labeling.
                activity_file (str): Path to the activity file for activity-based labeling.
                correction (list[str], optional): Optional list to correct column names.

            Returns:
                list[list[str]]: The labeled interaction matrix.
            """
            rows = [row.replace("\t", "") for row in rows]
            if self.codes and correction:
                columns = [correction[cont] + ", " + column for cont, column in enumerate(columns)]
            columns = [""] + columns[:]  # Add an empty string at the start for residue names
            
            if activity_file:
                if not os.path.isfile(activity_file):
                    raise FileNotFoundError(f"The file '{activity_file}' does not exist.")
                
                # Read activity data into a dictionary
                data_dict = {}
                with open(activity_file, newline='') as csvfile:
                    csvreader = csv.reader(csvfile)
                    try:
                        next(csvreader)  # Skip header
                    except:
                        raise ValueError(f"The CSV file '{activity_file}' is missing a header.")
                    for key, value in csvreader:
                        data_dict[key] = str(round(float(value), 3))

                if not data_dict:
                    raise ValueError(f"The CSV file '{activity_file}' must contain at least one row of data.")

                # Update column names with activity data
                if correction:
                    for i in range(1, len(columns)):
                        drug_name = columns[i]
                        columns[i] = f"{drug_name} ({data_dict.get(correction[i-1], '0')})"
                else:
                    for i in range(1, len(columns)):
                        drug_name = columns[i]
                        columns[i] = f"{drug_name} ({data_dict.get(drug_name, '0')})"

            else:
                for i in range(1, len(columns)):
                    drug_name = columns[i]
                    columns[i] = f"{drug_name} (0)"

            # Insert headers into the matrix
            matrix.insert(0, columns)
            for i, row in enumerate(matrix[1:], start=1):
                row.insert(0, rows[i-1])

            return matrix

        def modify_cell(
                text: str, 
                interaction: str, 
                atoms: str,
                interaction_labels: list,
                ) -> str:
            """
            Updates the cell content by adding interaction type and involved atoms.

            Args:
                text (str): Current content of the matrix cell.
                interaction (str): Type of interaction occurring.
                atoms (str): Atoms involved in the interaction.
                interaction_labels (list): List of predefined interaction labels.

            Returns:
                str: Updated cell content with formatted interaction information.
            """
            # Create an interaction_map based on the global INTERACTION_LABELS list
            interaction_map = {label: str(index + 1) for index, label in enumerate(interaction_labels)}

            # Assign the interaction code based on the interaction_map, or default to the last value
            interaction_code = interaction_map.get(interaction, str(len(interaction_labels)))

            # If the text is empty, add the interaction with the provided atoms
            if text == "":
                return f"{interaction_code} {GROUP_DELIM}{atoms}{GROUP_DELIM}"

            # Split the cell content and remove empty parts, keeping existing interactions
            content = text.replace(DIFF_DELIM, "").split(GROUP_DELIM)[:-1]
            exists = False
            cell = ''

            # Check if the interaction already exists and add atoms to the corresponding interaction
            for index, segment in enumerate(content):
                if index % 2 == 0 and interaction_code == segment.strip():
                    # Delete repate entries
                    entries = content[index + 1].split(SAME_DELIM)
                    detected = False
                    for entry in entries:
                        if entry == atoms:
                            detected = True
                            break
                    if not detected:
                        content[index + 1] += f"{SAME_DELIM}{atoms}"
                    exists = True
                    break

            # Rebuild the cell content, preserving existing interactions
            cell = DIFF_DELIM.join(f"{content[i]}{GROUP_DELIM}{content[i+1]}{GROUP_DELIM}" for i in range(0, len(content), 2))
            
            # If the interaction didn't exist, append it at the end
            if not exists:
                cell += f"{DIFF_DELIM}{interaction_code} {GROUP_DELIM}{atoms}{GROUP_DELIM}"
            
            return cell

        def read_file(
                file_name: str
                ) -> list[str]:
            """
            Reads the specified file and returns its content as a list of lines.

            Supports reading both JSON and text-based interaction files.

            Args:
                file_name (str): The name of the file to read.

            Returns:
                list[str]: List of lines from the file (for text files) or parsed JSON data.

            Raises:
                FileNotFoundError: If the file does not exist.
                Exception: If an unexpected error occurs during reading.
            """
            try:
                with open(file_name, 'r') as file:
                    if file_name.split('.')[-1] == "json":
                        return json.load(file)
                    else:
                        return [line.strip() for line in file.readlines()]
            except FileNotFoundError:
                print(f"Error: The file '{file_name}' does not exist.")
                raise
            except Exception as e:
                print(f"Error: An unexpected error occurred while reading '{file_name}': {e}")
                raise

        def adjust_subunits(
                matrix: list[list[str]], 
                ) -> list[list[str]]:
            """
            Adjusts matrix entries to correctly reflect subunit information for residues.

            Removes duplicate atoms and ensures that subunit data is accurately represented.

            Args:
                matrix (list[list[str]]): The interaction matrix to modify.

            Returns:
                list[list[str]]: Updated matrix with subunit adjustments.
            """
            
            def remove_duplicate_atoms(
                    atoms: str
                    ) -> str:
                """
                Removes duplicate atoms from the input string.

                Args:
                    atoms (str): A string containing atom interactions separated by SAME_DELIM.

                Returns:
                    str: A formatted string with unique atoms, enclosed in '|' delimiters.
                """
                # Split the input string by SAME_DELIM to get individual atoms.
                sections = atoms.split(SAME_DELIM)

                # Use a set to keep only unique atoms in their original order.
                unique_atoms = list(dict.fromkeys(sections))

                # Join the unique atoms back into a single string separated by SAME_DELIM.
                text = SAME_DELIM.join(unique_atoms)

                # Return the formatted string enclosed in '|' delimiters.
                return f"{GROUP_DELIM}{text}{GROUP_DELIM}"

            for row in range(len(matrix)):
                for column in range(len(matrix[row])):
                    cell = matrix[row][column]
                    if cell != '':
                        sections = cell.split(GROUP_DELIM)
                        text = ''.join(
                            sections[i-1] + remove_duplicate_atoms(sections[i])
                            for i in range(1, len(sections), 2)
                        )
                        matrix[row][column] = text
            return matrix

        def sort_interactions(
                matrix: list[list[str]]
                ) -> list[list[str]]:
            """
            Sorts the interaction types within each cell of the matrix in ascending order.

            Args:
                matrix (list[list[str]]): The matrix to be sorted.

            Returns:
                list[list[str]]: The sorted matrix with interactions ordered numerically.
            """
            for row_index, row in enumerate(matrix):
                for cell_index, cell in enumerate(row):
                    if cell != '':
                        # Split the cell into individual interactions
                        interactions = cell.split(DIFF_DELIM)

                        # Sort the interactions based on the number that follows the initial space
                        if len(interactions) > 1:
                            interactions = sorted(interactions, key=lambda x: int(x.split(" ")[0]))

                        # Join the sorted interactions back and update the cell
                        matrix[row_index][cell_index] = DIFF_DELIM.join(interactions)

            return matrix

        def validate_string(
                input_string: str
                ) -> bool:
            """
            Validates a residue string to ensure it follows the expected format.

            Format:
            - Starts with a three-letter amino acid abbreviation.
            - Followed by spaces and a numeric sequence.
            - Ends with a dash and additional characters.

            Args:
                input_string (str): The residue string to validate.

            Returns:
                bool: True if the format is valid, False otherwise.
            """
            # Regular expression to validate the required format
            pattern = r'^[A-Z]{3} +\d+-.+$'
            
            # Match the input string with the regular expression pattern
            return bool(re.match(pattern, input_string))

        def get_protein_ligand(begin: dict, end: dict) -> tuple[dict, dict]:
            if begin["label_comp_type"] == "P" and end["label_comp_type"] == "P":
                if begin["label_comp_id"] in self.aa and end["label_comp_id"] in self.aa:
                    return None, None
                elif begin["label_comp_id"] in self.aa:
                    return begin, end
                else:
                    return end, begin
            elif begin["label_comp_type"] == "P":
                return begin, end
            elif end["label_comp_type"] == "P":
                return end, begin
            else: 
                return None, None

        def validate_file(filename, mode):
            """
            Checks if a filename is valid (i.e., contains no spaces).

            Args:
                filename (str): Name of the file.

            Returns:
                bool: True if the filename is valid, False otherwise.
            """
            if filename.count(' ') != 0:
                print(f"Warning: The filename '{filename}' contains spaces.")
                return False
            if mode == self.ICHEM:
                if filename.split('.')[-1] != 'txt':
                    print(f"Warning: The filename '{filename}' is not a valid IChem file.")
                    return False
            elif mode == self.ARPEGGIO:
                if filename.split('.')[-1] != 'json':
                    print(f"Warning: The filename '{filename}' is not a valid Arpeggio file.")
                    return False
            return True
        
        def check_directory(directory):
            """
            Validates and retrieves the list of files in the specified directory.

            Args:
                directory (str): Path to the directory.

            Returns:
                list[str]: List of filenames in the directory.

            Raises:
                FileNotFoundError: If the directory does not exist.
                EmptyDirectoryException: If the directory is empty.
            """
            if not os.path.exists(directory):
                raise FileOrDirectoryException(path=directory, error_type='not_found')
            elif not os.path.isdir(directory):
                raise FileOrDirectoryException(path=directory, error_type='not_found')
            else:
                files = os.listdir(directory)
                if not files:
                    raise FileOrDirectoryException(path=directory, error_type='empty')
            return files

        def ichem_analysis(content, index, files, subunits_set, cont, matrix, aa):
            """
            Processes IChem interaction files, extracting relevant data and updating the matrix.
            """
            for line in content:
                elements = line.split(GROUP_DELIM)
                if len(elements) == 10:
                    interaction = elements[0].strip().replace("\t", "")
                    residue = elements[3].strip().replace("\t", "")
                    if validate_string(residue):
                        if not subunit:
                            sections = residue.split("-")
                            residue = sections[0]
                            subunits_set.add(sections[1])

                        atoms = f"{elements[1].strip()}-{elements[4].strip()}" if protein and ligand else elements[1].strip() if protein else elements[4].strip() if ligand else ""
                        if not subunit:
                            atoms += f"({sections[1]})"

                        if residue not in aa:
                            aa[residue] = cont
                            cont += 1
                        column = aa[residue]

                        # Ensure matrix size and modify cell
                        if len(matrix) <= column:
                            matrix.append([""] * len(files))

                        matrix[column][index] = modify_cell(text=matrix[column][index], interaction=interaction, atoms=atoms, interaction_labels=INTERACTION_LABELS)
            return matrix, aa, cont, subunits_set

        def arpeggio_analysis(content, index, files, subunits_set, cont, matrix, aa):
            """
            Processes Arpeggio interaction files, extracting relevant data and updating the matrix.
            """
            # Filter to obtain entries with interacting_entities == INTER
            inter_set = [elem for elem in content if elem["interacting_entities"] in ARPEGGIO_INT_ENT]
            # Filter to obtain entries with the desired contact or type
            inter_set = [
                elem for elem in inter_set
                for contact in elem["contact"]
                if contact in ARPEGGIO_CONT or elem["type"] in ARPEGGIO_TYPE
            ]
            for inter in inter_set:
                if inter["type"] in ARPEGGIO_TYPE:
                    contact = [inter["type"]]
                else:
                    contact = [conta for conta in inter["contact"] if conta in ARPEGGIO_CONT]
                prot, lig = get_protein_ligand(begin=inter["bgn"], end=inter["end"])
                if prot is not None:
                    residue = prot["label_comp_id"] + " " + str(prot["auth_seq_id"])
                    prot_atom = prot["auth_atom_id"]
                    prot_subunit = prot["auth_asym_id"]
                    ligand_code = lig["label_comp_id"]
                    lig_atom = lig["auth_atom_id"]
                    
                    subunits_set.add(prot_subunit)
                    atoms = f"{prot_atom}-{lig_atom}" if protein and ligand else prot_atom if protein else lig_atom if ligand else ""

                    if subunit:
                        residue += "-" + prot_subunit
                    else:
                        atoms += f"({prot_subunit})"

                    if residue not in aa:
                        aa[residue] = cont
                        cont += 1
                    column = aa[residue]

                    # Ensure matrix size and modify cell
                    if len(matrix) <= column:
                        matrix.append([""] * len(files))

                    for interaction in contact:
                        matrix[column][index] = modify_cell(text=matrix[column][index], interaction=interaction, atoms=atoms, interaction_labels=ARPEGGIO_CONT+ARPEGGIO_TYPE)

            return matrix, ligand_code, aa, cont, subunits_set
        
        # Validate input types
        self._check_variable_types(
            variables=[directory, mode, activity_file, protein, ligand, subunit, save],
            expected_types=[str, str, (str, None.__class__), bool, bool, bool, (str, None.__class__)],
            variable_names=['directory', 'mode', 'activity_file', 'protein', 'ligand', 'subunit', 'save']
        )

        directory = os.path.join(self.input_directory, directory)
        if activity_file is not None:
            activity_file = os.path.join(self.input_directory, activity_file)

        # Check the directory and return its files
        files = check_directory(directory=directory)
        
        # Check if the mode is registered
        if mode not in PROGRAM_MODES:
            raise InvalidModeException(mode=mode, expected_values=PROGRAM_MODES)
        else:
            self.set_config(mode=mode)

        ligands = [None] * len(files)
        matrix = []
        aa = {}
        cont = 0
        subunits_set = set()
        failed_files = []
        
        # Analyze each file in the directory
        for index, file in enumerate(files):
            file_path = os.path.join(directory, file)

            if os.path.isfile(file_path) and validate_file(file, mode):
                content = read_file(file_path)
                if mode == self.ICHEM:
                    matrix, aa, cont, subunits_set = ichem_analysis(content=content, index=index, files=files, subunits_set=subunits_set, cont=cont, matrix=matrix, aa=aa)
                    ligands[index] = file.replace(".txt", "")
                elif mode == self.ARPEGGIO:
                    matrix, ligand_code, aa, cont, subunits_set = arpeggio_analysis(content=content, index=index, files=files, subunits_set=subunits_set, cont=cont, matrix=matrix, aa=aa)
                    files[index] = file.replace(".json", "")
                    ligands[index] = ligand_code
            
            else:
                failed_files.append(file_path)
                
        if mode == self.ICHEM:
            files = None       
                
        if len(failed_files) > 0:
            files = [f for f in files if f not in failed_files]
            ligands = [l for l in ligands if l is not None]
        
        if not subunit:
            matrix = adjust_subunits(matrix=matrix)
        matrix = sort_interactions(matrix=matrix)
        matrix = label_matrix(matrix=matrix, rows=list(aa.keys()), columns=ligands, activity_file=activity_file, correction=files)
        interaction_data = InteractionData(colors=self.plot_colors, interactions=self.interaction_labels,
                                     ligand=ligand, matrix=matrix, mode=mode, protein=protein, subunit=subunit, subunits_set=subunits_set)
        interaction_data = self.sort_matrix(interaction_data=interaction_data, residue_chain=True)
        # Save the matrix if specified
        if save:
            self.save_interaction_data(interaction_data=interaction_data, filename=save)

        return interaction_data
    
    def filter_by_interaction(self, 
            interaction_data: InteractionData,
            interactions: list[int], 
            save: str = None
            ) -> InteractionData:
        """
        Filters an interaction matrix based on specified interaction types.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.
            interactions (list[int]): List of valid interaction types (numbers 1 to 7) to retain in the matrix.
            save (str, optional): File path to save the filtered matrix. Defaults to None.

        Returns:
            InteractionData: The updated InteractionData object with the filtered matrix.

        Raises:
            ValueError: If the matrix dimensions are invalid, or if no matching interactions are found.
        """

        def validate_list(interactions: list[int]) -> None:
            """
            Validates the interaction list to ensure it contains unique numbers between 1 and 7.

            Args:
                interactions (list): List of integers representing interaction types.

            Raises:
                ValueError: If any number is outside the range of 1 to 7 or if there are duplicates.
            """
            # Valid interactions are numbers from 1 to 7
            valid_numbers = set(range(1, len(self.interaction_labels) + 1))

            # Check if all numbers in the list are within the valid range
            for num in interactions:
                if num not in valid_numbers:
                    raise ValueError(f"Invalid interaction: {num}. Must be a number between 1 and 7.")

            # Ensure the list contains no duplicate values
            if len(set(interactions)) != len(interactions):
                raise ValueError("The interaction list must not contain duplicates.")

        # Validate types of the matrix, interactions, and save parameters
        self._check_variable_types(
            variables=[interaction_data, interactions, save], 
            expected_types=[InteractionData, list, (str, None.__class__)], 
            variable_names=['interaction_data', 'interactions', 'save']
        )

        data = copy.deepcopy(interaction_data)
        matrix = data.matrix

        # Validate that the matrix has appropriate dimensions
        self._verify_dimensions(matrix=matrix)

        # Validate that the interaction list contains valid values
        validate_list(interactions=interactions)

        # Track whether any interactions were filtered
        changes = False

        # Iterate through each cell in the matrix (skipping the header row/column)
        for i in range(1, len(matrix)):
            for j in range(1, len(matrix[i])):
                cell = matrix[i][j]
                
                # If the cell is not empty ('-'), process it
                if is_not_empty_or_dash(cell):
                    sections = cell.split(DIFF_DELIM)
                    cell = ""
                    
                    # Iterate through the sections in the cell to filter valid interactions
                    for section in sections:
                        # Check if the first number in the section is in the valid interaction list
                        if int(section.split(' ')[0]) in interactions:
                            # Add the section to the cell if it contains a valid interaction
                            if cell == '':
                                cell = section
                            else:
                                cell += DIFF_DELIM + section
                            changes = True
                    
                    # Update the cell with the filtered sections or set it to '-' if empty
                    matrix[i][j] = cell if cell != '' else EMPTY_DASH_CELL
        
        # If no changes were made, raise an error indicating no matching interactions were found
        if not changes:
            raise ValueError("No matching interactions were found in the matrix.")

        data.matrix = matrix

        # Save the filtered matrix to a file if a save path is provided
        if save:
            self.save_interaction_data(interaction_data=data, filename=save)

        return data

    def filter_by_subunit(self,
            interaction_data: InteractionData, 
            subunits: list[str], 
            save: str = None
            ) -> InteractionData:
        """
        Filters an interaction matrix based on specified subunits.

        This method processes an interaction matrix and removes rows or interaction elements 
        that do not match the provided subunit list. The filtering mechanism depends on whether 
        the matrix is organized by residues or interactions.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.
            subunits (list[str]): List of valid subunits used as filtering criteria.
            save (str, optional): File path to save the filtered matrix. Defaults to None.

        Returns:
            InteractionData: The updated InteractionData object with the filtered matrix.

        Raises:
            ValueError: If the matrix dimensions are invalid or if no matching subunits are found.
        """
        
        def get_subunits_location(matrix: list[list[str]]) -> str:
            """
            Determines the location of subunits in the matrix (residues or interactions).

            Args:
                matrix (list): The matrix being analyzed.

            Returns:
                str: 'residues' if the first column indicates residue data, otherwise 'interactions'.
            """
            return 'residues' if len(matrix[1][0].split('-')) == 2 else 'interactions'
        
        # Check types of the matrix, subunits, and save parameters
        self._check_variable_types(
            variables=[interaction_data, subunits, save], 
            expected_types=[InteractionData, list, (str, None.__class__)], 
            variable_names=['interaction_data', 'subunits', 'save']
        )

        data = copy.deepcopy(interaction_data)
        matrix = data.matrix

        # Validate the dimensions of the matrix
        self._verify_dimensions(matrix=matrix)

        # Determine the axis of residues in the matrix
        axis = self._get_residues_axis(matrix=matrix)

        # Transpose the matrix if the axis is columns
        if axis == "columns":
            matrix = self.transpose_matrix(matrix)
        
        # Determine whether the matrix contains residues or interactions
        subunitsLocation = get_subunits_location(matrix=matrix)

        # Initialize change tracking variables
        changes = 0

        # Filter based on residue locations
        if subunitsLocation == 'residues':
            for index in range(1, len(matrix)):
                sections = matrix[index - changes][0].split("-")
                # Remove rows without valid sections or subunits
                if len(sections) != 2 or sections[1] not in subunits:
                    matrix.pop(index - changes)
                    changes += 1
        else:
            # Iterate through each cell in the matrix to filter interactions
            for i in range(1, len(matrix)):
                for j in range(1, len(matrix[i])):
                    cell = matrix[i][j]
                    
                    # Process non-empty cells
                    if is_not_empty_or_dash(cell=cell):
                        sections = cell.split(DIFF_DELIM)
                        cell = ""
                        
                        # Iterate through each section in the cell
                        for section in sections:
                            separators = section.split(GROUP_DELIM)[:-1]

                            # Filter out unwanted interactions
                            for index in range(1, len(separators)):
                                if index % 2 != 0:
                                    interactions = separators[index].split(SAME_DELIM)
                                    subchanges = 0
                                    
                                    # Remove interactions not in the valid subunits
                                    for interaction in range(len(interactions)):
                                        if len(interactions[interaction-subchanges].split('(')) > 1 and interactions[interaction - subchanges][-2] not in subunits:
                                            changes += 1
                                            interactions.pop(interaction - subchanges)
                                            subchanges += 1

                                    # Rebuild the cell if there are valid interactions
                                    if interactions:
                                        cell += separators[index - 1] + GROUP_DELIM
                                        cell += SAME_DELIM.join(interactions) + ' ' + GROUP_DELIM + DIFF_DELIM
                                        
                        # Update the cell with filtered interactions
                        cell = cell[:-2]  # Remove trailing comma and space
                        matrix[i][j] = cell if cell else EMPTY_DASH_CELL

        # Transpose the matrix back if it was originally in columns
        if axis == "columns":
            matrix = self.transpose_matrix(matrix)

        data.matrix = matrix

        # Save the filtered matrix to a file if a save path is provided
        if save:
            self.save_interaction_data(interaction_data=data, filename=save)

        return data

    def filter_by_chain(self, 
        interaction_data: InteractionData, 
        chain: str = None, 
        subpocket_path: str = None, 
        subpockets: list[str] = None, 
        save: str = None
    ) -> InteractionData:
        """
        Filters an interaction matrix based on a specified chain or subpockets.

        This method processes an interaction matrix and removes rows or interaction elements 
        that do not match the provided chain or residues extracted from the given subpockets.

        The filtering can be performed in two ways:
        1. By specifying a `chain` ("<main>" or "<side>"), which filters interactions 
        based on the presence of main or side chain atoms.
        2. By providing a `subpocket_path` and a list of `subpockets`, which extracts 
        residues from a predefined subpocket file and filters interactions accordingly.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.
            chain (str, optional): Specifies whether to retain "<main>" or "<side>" interactions. Defaults to None.
            subpocket_path (str, optional): Path to the file containing subpocket residue definitions. Defaults to None.
            subpockets (list[str], optional): List of subpockets to use for residue-based filtering. Defaults to None.
            save (str, optional): File path to save the filtered matrix. Defaults to None.

        Returns:
            InteractionData: The updated InteractionData object with the filtered matrix.

        Raises:
            ValueError: If the matrix dimensions are invalid.
            InvalidModeException: If an invalid chain mode is specified.
            Exception: If no protein atoms are available in the interaction data when filtering by chain.
        """

        def extract_subpockets_from_file(subpocket_file_path: str, subpocket_list: list[str]) -> list[str]:
            """Extract residues from the subpocket file matching specified subpockets."""
            residues = []
            with open(subpocket_file_path, mode='r', encoding='utf-8') as csv_file:
                reader = csv.reader(csv_file)
                for row in reader:
                    if row[0] in subpocket_list:
                        second_column_items = row[1].split(', ')
                        for item in second_column_items:
                            split_values = item.split('<')
                            if len(split_values) == 2:
                                validate_chain("<" + split_values[1])  # Validate chain format
                            residues.append(item)
            return residues
        
        def filter_row(row: list[str], chain_to_filter: str) -> tuple[list[str], bool]:
            """
            Filters a single row of the matrix based on chain criteria.
            
            Args:
                row (list[str]): A row from the interaction matrix.
                chain_to_filter (str): The chain to filter.

            Returns:
                tuple: Filtered row and a boolean indicating if the row is empty.
            """
            atom_types = ["C", "CA", "N", "O"]
            atom_filter = (
                lambda atom: atom in atom_types
                if chain_to_filter == "<main>"
                else atom not in atom_types
            )

            filtered_row = [row[0]]  # Keep the first column (residue)
            row_is_empty = True

            for cell in row[1:]:
                if is_not_empty_or_dash(cell):
                    new_cell_content = ""
                    for interaction in cell.split(DIFF_DELIM):
                        interaction_number, interaction_pairs = interaction.split(GROUP_DELIM, 1)
                        valid_pairs = [
                            pair for pair in interaction_pairs.split(SAME_DELIM)
                            if any(atom_filter(atom) for atom in pair.split("-")[0].split(","))
                        ]
                        if valid_pairs:
                            new_cell_content += f"{interaction_number}{GROUP_DELIM}{SAME_DELIM.join(valid_pairs)}{DIFF_DELIM}"
                    new_cell_content = new_cell_content[:-len(DIFF_DELIM)] if new_cell_content else "-"
                    filtered_row.append(new_cell_content)
                    if new_cell_content != "-":
                        row_is_empty = False
                else:
                    filtered_row.append(cell)

            return filtered_row, row_is_empty

        def filter_matrix(matrix: list[list[str]], residues: list[str], chain: str) -> list[list[str]]:
            """
            Filters the interaction matrix based on residues and chain.

            Args:
                matrix (list[list[str]]): The original interaction matrix.
                residues_to_filter (list[str]): List of residues to filter.
                chain (str): The chain to filter.

            Returns:
                list[list[str]]: The filtered interaction matrix.
            """
            filtered_matrix = [matrix[0]]  # Keep the header row

            # Prepare comparable dictionary for quick lookups
            residue_chain_map = {
                item.split("<")[0]: f"<{item.split('<')[1]}" if len(item.split("<")) == 2 else "<all>"
                for item in residues
            }

            for row in matrix[1:]:
                residue = row[0].replace(" ", "").split("-")[0]
                if residue in residue_chain_map:
                    if residue_chain_map[residue] == "<all>":
                        filtered_matrix.append(row)
                    else:
                        filtered_row, is_empty = filter_row(row, residue_chain_map[residue])
                        if not is_empty:
                            filtered_matrix.append(filtered_row)
                elif chain:
                    filtered_row, is_empty = filter_row(row, chain)
                    if not is_empty:
                        filtered_matrix.append(filtered_row)

            return filtered_matrix

        def validate_chain(chain: str) -> None:
            """Validates that the chain is either '<main>' or '<side>'."""
            valid_chains = ["<main>", "<side>"]
            if chain not in valid_chains:
                raise InvalidModeException(mode=chain, expected_values=valid_chains)

        # Check types of the matrix, chain, and subpocket
        self._check_variable_types(
            variables=[interaction_data, chain, subpockets, subpocket_path, save],
            expected_types=[InteractionData, (str, None.__class__), (list, None.__class__), (str, None.__class__), (str, None.__class__)],
            variable_names=['interaction_data', 'chain', 'subpockets', 'subpocket_path', 'save']
        )

        filtered_data = copy.deepcopy(interaction_data)
        matrix = filtered_data.matrix
        residues_selection = []

        # Extract residues if subpockets are provided
        if subpocket_path and subpockets:
            subpocket_path = os.path.join(self.input_directory, subpocket_path)
            residues = extract_subpockets_from_file(subpocket_path, subpockets)
            if chain:
                validate_chain(chain=chain)
                if not filtered_data.protein:
                    raise Exception("No protein atoms available in the interaction data.")
                for residue in residues:
                    values = residue.split('<')
                    if len(values) == 2:
                        if "<" + values[1] == chain:
                            residues_selection.append(residue)
                    else:
                        residues_selection.append(residue + chain)
                chain = None
            else:
                residues_selection = residues
        elif chain:
            validate_chain(chain=chain)
            if not filtered_data.protein:
                raise Exception("No protein atoms available in the interaction data.")

        # Validate the dimensions of the matrix
        self._verify_dimensions(matrix=matrix)

        axis = self._get_residues_axis(matrix=matrix)

        if axis == 'columns':
            matrix = self.transpose_matrix(matrix)

        filtered_matrix = filter_matrix(matrix=matrix, residues=residues_selection, chain=chain)

        if axis == 'columns':
            filtered_matrix = self.transpose_matrix(filtered_matrix)

        filtered_data.matrix = filtered_matrix

        if save:
            self.save_interaction_data(interaction_data=filtered_data, filename=save)

        return filtered_data

    def get_dataframe(self, interaction_data: InteractionData) -> pd.DataFrame:
        """
        Converts the interaction matrix into a pandas DataFrame.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.

        Returns:
            pd.DataFrame: A DataFrame representation of the interaction matrix.
        """
        # Validate the type of interaction_data
        self._check_variable_types(
            variables=[interaction_data],
            expected_types=[InteractionData],
            variable_names=['interaction_data']
        )

        # Convert the matrix to a DataFrame and set the index and columns
        df = pd.DataFrame(interaction_data.matrix[1:], columns=interaction_data.matrix[0])
        df.set_index(df.columns[0], inplace=True)

        return df

    def get_interactions(self, interaction_data: InteractionData) -> list[str]:
        """
        Retrieves the interaction labels from the InteractionData object.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.

        Returns:
            list[str]: A list of interaction labels.
        """
        # Validate the type of interaction_data
        self._check_variable_types(
            variables=[interaction_data],
            expected_types=[InteractionData],
            variable_names=['interaction_data']
        )

        return interaction_data.interactions
    
    def heatmap(self, interaction_data: InteractionData, title: str, mode: str, x_label: str = "", y_label: str = "", min_v: int = None, max_v: int = None, font: str = None, save: bool = False):
        """
        Generates a heatmap based on interaction data using different processing modes.

        This method processes the interaction matrix, computes interaction statistics 
        according to the specified mode, and generates a heatmap visualization. If the 
        dataset is large, the visualization is split into multiple heatmaps.

        Supported processing modes:
            - 'min': Displays the minimum interaction values.
            - 'max': Displays the maximum interaction values.
            - 'mean': Computes and visualizes the average interaction values.
            - 'count': Counts the occurrences of interactions.
            - 'percent': Displays the percentage of interactions.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.
            title (str): Title of the heatmap.
            mode (str): Processing mode ('min', 'max', 'mean', 'count', or 'percent').
            x_label (str, optional): Label for the x-axis. Defaults to an empty string.
            y_label (str, optional): Label for the y-axis. Defaults to an empty string.
            min_v (int, optional): Minimum value for the heatmap color scale. Defaults to None (auto-scaling).
            max_v (int, optional): Maximum value for the heatmap color scale. Defaults to None (auto-scaling).
            font (str, optional): Font style for the heatmap labels. Can be 'upper', 'lower', or None. Defaults to None.
            save (bool, optional): If True, saves the heatmap instead of displaying it. Defaults to False.

        Returns:
            None: The function either displays the heatmap(s) or saves them to files.

        Raises:
            InvalidModeException: If an unsupported processing mode is provided.
            HeatmapActivityException: If the matrix contains invalid or negative activity values.
        """

        def validate_and_prepare_matrix(matrix: list[list[str]]):
            """
            Validates the input matrix and prepares it for processing.
            
            This method checks that all ligand activity values in the matrix are non-negative and transposes the matrix 
            if the residue axis is configured to be in columns.

            Args:
                matrix (list[list[str]]): The matrix to validate and prepare.

            Returns:
                list[list[str]]: The validated matrix.

            Raises:
                HeatmapActivityException: If any ligand activity value is negative.
            """
            if self._get_residues_axis == 'columns':
                matrix = self.transpose_matrix(matrix)
            for ligand in matrix[0][1:]:
                activity = ligand.split('(')[-1].replace(")", "")
                if float(activity) < 0.0:
                    raise HeatmapActivityException
            return matrix

        def process_matrix(matrix: list[list[str]], mode: str) -> dict:
            """
            Processes the matrix based on the specified mode.
            
            Depending on the mode, calculates metrics such as minimum, maximum, mean, count, or percentage of interactions.

            Args:
                matrix (list[list[str]]): The input matrix to process.
                mode (str): The mode to process the data. Supported modes are 'min', 'max', 'mean', 'count', and 'percent'.

            Returns:
                dict: A dictionary mapping residues to interaction values processed according to the mode.
            """
            data, matrix = initialize_data(matrix=matrix)
            op = {"min": operator.lt, "max": operator.gt}.get(mode, None)

            for line in matrix[1:]:
                activity = float(line[0].split('(')[-1].replace(")", ""))
                for index in range(1, len(line)):
                    residue = matrix[0][index].split('-')[0]
                    cell = line[index]
                    sections = cell.split(DIFF_DELIM)

                    for section in sections:
                        if is_not_empty_or_dash(section.split(' ')[0]):
                            interaction = int(section.split(' ')[0]) - 1
                            current_value = data[residue][interaction]

                            if mode in ["min", "max"]:
                                if np.isnan(current_value) or op(activity, current_value):
                                    data[residue][interaction] = activity
                            elif mode == "mean":
                                if not isinstance(current_value, list):
                                    data[residue][interaction] = [1, activity]
                                else:
                                    count, total = current_value
                                    data[residue][interaction] = [count + 1, total + activity]
                            elif mode in ["count", "percent"]:
                                if np.isnan(current_value):
                                    data[residue][interaction] = 1
                                else:
                                    data[residue][interaction] += 1

            if mode == "mean":
                for residue, interactions in data.items():
                    for i, value in enumerate(interactions):
                        if isinstance(value, list):
                            count, total = value
                            data[residue][i] = round(total / count, 2)
            elif mode == "percent":
                num_lines = len(matrix) - 1
                for residue, interactions in data.items():
                    for i, value in enumerate(interactions):
                        data[residue][i] = (value / num_lines * 100) if not np.isnan(value) else np.nan

            return data

        def initialize_data(matrix: list[list[str]]) -> list[list[float]]:
            """
            Initializes an empty data structure to store processed interaction values.

            Args:
                matrix (list[list[str]]): The input matrix to initialize data for.

            Returns:
                tuple: A tuple containing:
                    - dict: The initialized data structure with residues as keys and NaN interaction values.
                    - list[list[str]]: The transposed matrix for processing.
            """
            data = {}
            for line in matrix[1:]:
                residue = line[0].split('-')[0]
                data[residue] = [np.nan for _ in range(len(self.interaction_labels))]
            return data, self.transpose_matrix(interaction_data=matrix)

        def plot_heatmap(self, data, title, x_label, y_label, mode, min_v, max_v, save, font):
            """
            Creates and optionally saves the heatmap visualization.

            Args:
                data (dict): The processed interaction data to visualize.
                title (str): The title for the heatmap.
                x_label (str): Label for the x-axis.
                y_label (str): Label for the y-axis.
                mode (str): The mode used to process the data.
                min_v (int, optional): Minimum value for the heatmap color scale.
                max_v (int, optional): Maximum value for the heatmap color scale.
                save (bool): Whether to save the heatmap to a file.

            Returns:
                None: Displays the heatmap or saves it to a file.
            """
            df = pd.DataFrame(data, index=self.interaction_labels if font == None else [elemento.upper() for elemento in self.interaction_labels] if font == "upper" else [elemento.lower() for elemento in self.interaction_labels])
            max_cols = self.heat_max_cols  # Maximum number of columns per heatmap
            num_cols = len(df.columns)

            vmin = min_v if min_v else df.min().min()
            vmax = max_v if max_v else df.max().max()

            num_heatmaps = (num_cols + max_cols - 1) // max_cols  # Round up
            cols_per_heatmap = (num_cols + num_heatmaps - 1) // num_heatmaps  # Distribute evenly

            if mode == 'count':
                tagMin = f"{vmin:.0f}"
                tagMax = f"{vmax:.0f}"
            else:
                tagMin = f"{vmin:.1f}"
                tagMax = f"{vmax:.1f}"

            # if title ends with '(/)', it will not be displayed in the heatmap
            display = True
            if title.endswith("(/)"):
                title = title[:-3]
                display = False

            for i in range(num_heatmaps):
                start_col = i * cols_per_heatmap
                end_col = min((i + 1) * cols_per_heatmap, num_cols)
                df_subset = df.iloc[:, start_col:end_col]

                plt.figure(figsize=(14, 9))
                sns.heatmap(df_subset, 
                            annot=True,
                            linewidths=0.5, 
                            linecolor='lightgrey',
                            cmap=self.heat_colors, 
                            fmt=".0f" if mode == 'count' else ".1f", 
                            vmin=vmin, 
                            vmax=vmax, 
                            cbar_kws={
                                "ticks": np.linspace(vmin, vmax, num=6),
                                "format": "%.0f" if mode == 'count' else "%.1f"
                            })

                # if title ends with '(/)', it will not be displayed in the heatmap
                if display:
                    if num_heatmaps == 1:
                        plt.title(f"{title}")
                    else:
                        plt.title(f"{title} (Columns {start_col + 1}-{end_col})")

                plt.xlabel(x_label)
                plt.ylabel(y_label)
                plt.subplots_adjust(left=0.145, bottom=0.155, right=1, top=0.935)

                if not save:
                    plt.show()
                else:
                    filename = os.path.join(self.saving_directory, f"{title}_part_{i + 1}.png")
                    plt.savefig(filename)
                    plt.close()

        self._verify_font(font=font)

        data = copy.deepcopy(interaction_data)
        matrix = data.matrix

        self.set_config(interaction_data=interaction_data)

        # Validate and prepare the matrix
        matrix = validate_and_prepare_matrix(matrix=matrix)

        # Ensure the mode is valid
        if mode not in HEATMAP_MODES:
            raise InvalidModeException(mode=mode, expected_values=HEATMAP_MODES)

        # Process the matrix based on the mode
        data = process_matrix(matrix=matrix, mode=mode)

        # Generate and display/save the heatmap
        plot_heatmap(self=self, data=data, title=title, x_label=x_label, y_label=y_label, mode=mode, min_v=min_v, max_v=max_v, save=save, font=font)

    def bar_chart(self,
        interaction_data: InteractionData,
        plot_name: str,
        axis: str,
        label_x: str = None,
        label_y: str = "Number of intermolecular interactions",
        title: str = "Protein-drug interactions",
        stacked: bool = False,
        save: bool = False,
        colors: list[str] = None,
        type_count: bool = False,
        font: str = None
    ) -> None:
        """
        Generates a bar chart based on interaction data.

        This method extracts relevant data from an interaction matrix and visualizes it 
        as a bar chart. If the dataset is too large, it automatically splits the data into 
        multiple plots for better readability.

        The method supports:
        - **Standard bar charts**: A single bar per residue or PDB complex.
        - **Stacked bar charts** (`stacked=True`): Bars grouped by interaction types, 
        showing their relative contribution.
        - **Interactive annotations**: Hovering over stacked bars displays the percentage 
        of each interaction type.
        - **Automatic data splitting**: Large datasets are split into multiple charts.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.
            plot_name (str): Name of the plot (used for saving).
            axis (str): Defines whether to plot rows ('rows') or columns ('columns').
            label_x (str, optional): Label for the x-axis. Defaults to "Interacting protein residues".
            label_y (str, optional): Label for the y-axis. Defaults to "Number of intermolecular interactions".
            title (str, optional): Title of the chart. Defaults to "Protein-drug interactions".
            stacked (bool, optional): If True, creates a stacked bar chart. Defaults to False.
            save (bool, optional): If True, saves the chart as a PNG file. Defaults to False.
            colors (list[str], optional): List of colors for interaction types. Defaults to None.
            type_count (bool, optional): If True, counts the occurrences of each interaction type. Defaults to False.
            font (str, optional): Font style for the plot leyend. Can be 'upper', 'lower', or None. Defaults to None.

        Returns:
            None: The function either displays or saves the plot.

        Raises:
            ValueError: If `axis` is not 'rows' or 'columns'.
        """

        self._verify_font(font=font)

        matrix = interaction_data.matrix
        self.set_config(interaction_data=interaction_data)
        # Initialize and get data
        colors, data, indices, transposed_data = self._plot_init(colors, matrix, axis, type_count)
        max_elements_plot = self.plot_max_cols
        num_x_elements = len(indices)

        # Divide data into multiple plots if necessary
        if num_x_elements > max_elements_plot:
            # Calculate the number of plots needed
            num_plots = (num_x_elements + max_elements_plot - 1) // max_elements_plot  # Ceiling division

            # Calculate the number of elements per plot, ensuring equal distribution
            elements_per_plot = num_x_elements // num_plots
            remainder = num_x_elements % num_plots

            # Split indices into equally distributed groups
            split_indices = []
            start = 0
            for i in range(num_plots):
                # Distribute the remainder among the first few plots
                end = start + elements_per_plot + (1 if i < remainder else 0)
                split_indices.append((start, end))
                start = end

            # Ensure all plots have the same Y-axis limit
            max_y_values = []
            split_data = []

            for start, end in split_indices:
                if stacked:
                    # Safely slice transposed_data based on start and end
                    split_group = [group[start:end] for group in transposed_data]
                    split_data.append(split_group)
                    max_y_values.append(max(sum(col) for col in zip(*split_group)))
                else:
                    # Safely slice data based on start and end
                    split_data.append((data[start:end]))
                    max_y_values.append(max(split_data[-1][1]))

            global_max_y = max(max_y_values)

            # Generate separate plots for each split group
            for i, (start, end) in enumerate(split_indices):
                fig, ax = plt.subplots(num=f"{plot_name}_{i+1}", figsize=(12, 6))
                subset_indices = indices[start:end]  # Subset of indices for the current plot

                if stacked:
                    bars = []
                    bottoms = [0] * (end - start)
                    for index, group in enumerate(transposed_data):
                        group_subset = group[start:end]  # Subset of the group for the current plot
                        bar = ax.bar(subset_indices, group_subset, bottom=bottoms, label=self.interaction_labels[index], color=colors[index])
                        bars.append(bar)
                        bottoms = [b + g for b, g in zip(bottoms, group_subset)]

                    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=1)

                    # Add interactive cursors to display percentages for stacked bars
                    cursor = mplcursors.cursor(bars, hover=True)

                    @cursor.connect("add")
                    def on_add(sel):
                        # Get the index of the bar in the current figure
                        bar_index = sel.index
                        # Translate the index to the original dataset's range
                        original_index = start + bar_index
                        # Calculate totals and percentages
                        total = sum(transposed_data[j][original_index] for j in range(len(transposed_data)))
                        percentages = [
                            transposed_data[j][original_index] / total * 100 if total != 0 else 0
                            for j in range(len(transposed_data))
                        ]
                        # Create annotation text
                        annotation_text = "\n".join(
                            [f"{self.interaction_labels[j]}: {percentages[j]:.1f}%" for j in range(len(self.interaction_labels))]
                        )
                        sel.annotation.set_text(annotation_text)
                        sel.annotation.get_bbox_patch().set(fc="white", alpha=0.9)  # Set background to white with 90% opacity
                else:
                    x_data = data[0][start:end]
                    y_data = data[1][start:end]
                    ax.bar(x_data, y_data, color=colors[0] if len(colors) > 0 else None)

                ax.set_ylim(0, global_max_y * 1.1)  # Same Y-axis limit for consistency
                ax.set_xticks(range(len(subset_indices)))
                ax.set_xticklabels(subset_indices, rotation=90, ha='center')
                ax.set_ylabel(label_y)
                ax.set_xlabel(label_x or "Interacting protein residues")
                ax.set_title(f"{title} (Part {i+1})")
                ax.yaxis.set_major_locator(MaxNLocator(integer=True))
                plt.tight_layout()

                self._plot_end(save, plt, fig, f"{plot_name}_part_{i+1}")

        else:
            # Original plotting logic if data fits in one plot
            fig, ax = plt.subplots(num=plot_name, figsize=(12, 6))
            if stacked:
                bars = []
                bottoms = [0] * len(indices)
                for index, group in enumerate(transposed_data):
                    label = self.interaction_labels[index] if font is None else self.interaction_labels[index].upper() if font == "upper" else self.interaction_labels[index].lower()
                    bars.append(ax.bar(indices, group, bottom=bottoms, label=label, color=colors[index]))
                    bottoms = [i + j for i, j in zip(bottoms, group)]

                ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=1)
                max_y = max([sum(col) for col in data])
            else:
                if type_count:
                    data = self.sort_matrix(interaction_data, axis, count=type_count)
                    ax.bar(data[0], data[1], color=colors[0] if len(colors) > 0 else None)
                    max_y = max(data[1])
                else:
                    for i in range(len(data)):
                        data[i] = sum(data[i])
                    ax.bar(indices, data, color=colors[0] if len(colors) > 0 else None)
                    max_y = max(data)

            ax.set_ylim(0, max_y * 1.1)
            ax.set_xticks(range(len(indices)))
            ax.set_xticklabels(indices, rotation=90, ha='center')
            ax.set_ylabel(label_y)
            ax.set_xlabel(label_x or "Interacting protein residues")
            ax.set_title(title)
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            plt.tight_layout()

        # Add interactive cursors to display percentages for stacked bars
        if stacked:
            cursor = mplcursors.cursor(bars, hover=True)

            @cursor.connect("add")
            def on_add(sel):
                index = sel.index
                total = sum(transposed_data[i][index] for i in range(len(transposed_data)))
                percentages = [transposed_data[i][index] / total * 100 if total != 0 else 0 for i in range(len(transposed_data))]
                annotation_text = "\n".join([f"{self.interaction_labels[i]}: {percentages[i]:.1f}%" for i in range(len(self.interaction_labels))])
                sel.annotation.set_text(annotation_text)
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.9)  # Set background to white with 90% opacity

        self._plot_end(save, plt, fig, plot_name)

    def pie_chart(self,
        interaction_data: InteractionData,
        plot_name: str,
        axis: str,
        save: bool = False,
        colors: list[str] = None,
        type_count: bool = False,
        font: str = None
    ) -> None:
        """
        Generates a pie chart based on interaction data.

        This method processes an interaction matrix, calculates the proportion of different interaction 
        types, and visualizes them as a pie chart. Interaction types with zero occurrences are automatically 
        filtered out. The chart can be displayed or saved as a PNG file.

        Args:
            interaction_data (InteractionData): The object containing the interaction matrix.
            plot_name (str): The name of the plot (used for saving).
            axis (str): Defines whether to analyze rows ('rows') or columns ('columns').
            save (bool, optional): If True, saves the pie chart instead of displaying it. Defaults to False.
            colors (list[str], optional): List of colors for interaction types. Defaults to None.
            type_count (bool, optional): If True, counts the occurrences of each interaction type instead of 
                                        using interaction values. Defaults to False.
            font (str, optional): Font style for the plot leyend. Can be 'upper', 'lower', or None. Defaults to None.

        Returns:
            None: The function either displays or saves the plot.
        """

        def _filter_non_zero_interactions(total_interactions: list[int], colors: list[str]) -> list[tuple]:
            """
            Filters out interaction types with zero total count.

            Args:
                total_interactions (list[int]): Total interactions per type.
                colors (list[str]): List of colors for each interaction type.

            Returns:
                list[tuple]: Filtered interaction data as (label, total, color).
            """
            return[(label, total, colors[i])
                for i, (label, total) in enumerate(zip(self.interaction_labels, total_interactions))
                if total > 0]

        def _plot_pie_chart(labels: list[str], sizes: list[int], colors: list[str], total_interactions: list[int], font: str) -> None:
            """
            Creates and formats a pie chart.

            Args:
                labels (list[str]): Interaction labels for the pie chart.
                sizes (list[int]): Interaction sizes (counts or percentages).
                colors (list[str]): Colors for each interaction type.
                total_interactions (list[int]): Total counts of interactions per type.

            Returns:
                None
            """
            fig, ax_pie = plt.subplots(figsize=(10, 6))

            # Plot the pie chart
            ax_pie.pie(sizes, labels=None, colors=colors, autopct='', startangle=140)
            ax_pie.set_title('Interaction Percentages')

            # Calculate percentages and create legend
            total = sum(total_interactions)
            legend_labels = [
                f"{label if font is None else label.upper() if font == 'upper' else label.lower()} ({round(count / total * 100, 2)}%)"
                for label, count in zip(self.interaction_labels, total_interactions)
                if count != 0
            ]
            ax_pie.legend(legend_labels, title="Interaction Types", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

            return fig

        def _prepare_pie_chart_data(non_zero_interactions: list[tuple]) -> tuple:
            """
            Prepares data for plotting the pie chart.

            Args:
                non_zero_interactions (list[tuple]): Interaction data with non-zero totals.

            Returns:
                tuple: Labels, sizes, and colors for the pie chart.
            """
            if non_zero_interactions:
                return zip(*non_zero_interactions)  # Separates into labels, sizes, and colors
            return [], [], []
        
        self._verify_font(font=font)
        self.set_config(interaction_data=interaction_data)
        matrix = interaction_data.matrix
        # Initialize plotting parameters and data
        colors, data, indices, transposed_data = self._plot_init(colors, matrix, axis, type_count)

        # Calculate total interaction and prepade data for plotting
        total_interactions = [sum(transposed_data[i]) for i in range(len(transposed_data))]
        non_zero_interactions = _filter_non_zero_interactions(total_interactions, colors)
        labels_pie, sizes, pie_colors = _prepare_pie_chart_data(non_zero_interactions)
        
        # Plot the pie chart
        fig = _plot_pie_chart(labels_pie, sizes, pie_colors, total_interactions, font)

        # Show or save the plot
        self._plot_end(save, plt, fig, plot_name)

    def remove_empty_axis(
            self, 
            interaction_data: InteractionData,
            save: str = None
            ) -> InteractionData:
        """
        Removes empty rows and columns from the interaction matrix.

        This method iterates through the interaction matrix and removes rows and columns 
        that contain only empty or placeholder values (e.g., dashes or empty strings). 
        The cleaned matrix maintains the original structure and is optionally saved to a file.

        Args:
            interaction_data (InteractionData): The input interaction data from which empty rows and columns will be removed.
            save (str, optional): The filename to save the cleaned matrix. If None, the matrix will not be saved.

        Returns:
            InteractionData: The cleaned interaction data with empty rows and columns removed.

        Raises:
            TypeMismatchException: If the types of input variables do not match the expected types.
            ValueError: If the matrix dimensions are invalid or if it is too small after cleaning.
        """
        
        def _remove_empty_rows(matrix: list[list[str]]) -> list[list[str]]:
            """
            Helper function to remove empty rows from the matrix.

            Args:
                matrix (list[list[str]]): The matrix from which empty rows will be removed.

            Returns:
                list[list[str]]: The matrix with empty rows removed.
            """
            changes = 0
            for row in range(1, len(matrix)):
                if all(not is_not_empty_or_dash(cell=column) for column in matrix[row - changes][1:]):
                    matrix.pop(row - changes)
                    changes += 1
            return matrix

        self._check_variable_types(
            variables=[interaction_data, save], 
            expected_types=[InteractionData, (str, None.__class__)], 
            variable_names=['interaction_data', 'save']
        )

        data = copy.deepcopy(interaction_data)
        matrix = data.matrix

        self._verify_dimensions(matrix=matrix)

        matrix = _remove_empty_rows(matrix=matrix)
        
        # Transpose the matrix, remove empty columns (which are now rows)
        matrix = self.transpose_matrix(matrix)
        matrix = _remove_empty_rows(matrix=matrix)
        
        # Transpose back to restore original format
        matrix = self.transpose_matrix(matrix)

        data.matrix = matrix

        if save:
            self.save_interaction_data(interaction_data=data, filename=save)
        
        return data
    
    def save_interaction_data(
            self,
            interaction_data: InteractionData,
            filename: str
        ) -> None:
        """
        Saves the interaction data to an Excel file in the specified directory.

        This method exports the interaction matrix and additional attributes to an Excel file 
        with two separate sheets:
            - **Matrix**: Contains the interaction matrix.
            - **Attributes**: Stores metadata such as interaction types, colors, ligand status, 
            processing mode, protein consideration, and subunit status.

        Args:
            interaction_data (InteractionData): The interaction data to be saved.
            filename (str): The name of the output file (must end with '.xlsx').

        Returns:
            None

        Raises:
            ValueError: If the matrix dimensions are not valid.
            InvalidFileExtensionException: If the filename does not end with '.xlsx'.
            TypeMismatchException: If a variable type doesn't match the expected one.
        """

        # Validate variable types
        self._check_variable_types(
            variables=[interaction_data, filename],
            expected_types=[InteractionData, str],
            variable_names=['interaction_data', 'filename']
        )

        # Ensure the filename ends with .xlsx
        if not filename.endswith('.xlsx'):
            raise InvalidFileExtensionException(filename)

        # Validate matrix dimensions
        self._verify_dimensions(matrix=interaction_data.matrix)

        # Construct the file path
        file_path = os.path.join(self.saving_directory, filename)

        # Prepare data for the first sheet (Matrix)
        matrix_df = pd.DataFrame(interaction_data.matrix)

        # Prepare data for the second sheet (Attributes)
        # Create id, interacciones, colores columns for lists
        empty_strings = [""] * (len(interaction_data.interactions) - 1)  # Donde 'n' es el nmero de cadenas vacas que quieres.
        subunits_str = ", ".join(sorted(interaction_data.subunits_set))

        interactions_data = pd.DataFrame({
            "id": range(1, len(interaction_data.interactions) + 1),
            "interactions": interaction_data.interactions,
            "colors": interaction_data.colors,
            "ligand": ["True" if interaction_data.ligand else "False"] + empty_strings,
            "mode": [interaction_data.mode] + empty_strings,
            "protein": ["True" if interaction_data.protein else "False"] + empty_strings,
            "subunit": ["True" if interaction_data.subunit else "False" + f" ({subunits_str})"] + empty_strings
        })

        # Save to an Excel file with two sheets
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # First sheet: Matrix
            matrix_df.to_excel(writer, sheet_name='Matrix', index=False, header=False)

            # Second sheet: Interactions & Attributes
            # Write interactions and colors
            interactions_data.to_excel(writer, sheet_name='Attributes', index=False, startrow=0)

        # Apply background colors to the row "colors"
        wb = load_workbook(file_path)
        ws = wb["Attributes"]

        col_index = 3  # Column "colors" (1-based index in Excel)

        for row_idx, color_hex in enumerate(interaction_data.colors, start=2):  # Starts on row 2 (skipping the header)
            fill = PatternFill(start_color=color_hex.replace("#", ""), end_color=color_hex.replace("#", ""), fill_type="solid")
            ws.cell(row=row_idx, column=col_index).fill = fill

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Aadir un poco de espacio extra

        wb.save(file_path)

        print(f"Interaction data successfully saved to {file_path}")

    def sort_matrix(self,
        interaction_data: InteractionData,
        axis: str = 'rows',
        thr_interactions: int = None,
        thr_activity: float = None,
        selected_items: int = None,
        count: bool = False,
        residue_chain: bool = False,
        save: str = None
    ) -> InteractionData:
        """
        Sorts and filters rows or columns in the interaction matrix based on interaction criteria.

        This method sorts and selects reactive rows or columns in the interaction matrix according 
        to the specified criteria. It supports selection based on:
        - **Minimum interaction count** (`thr_interactions`): Keeps rows/columns with at least 
        the specified number of interactions.
        - **Activity threshold** (`thr_activity`): Selects rows/columns where the activity value 
        meets or exceeds the given threshold.
        - **Top-ranked selection** (`selected_items`): Retains only the top N rows/columns based 
        on interaction count.
        - **Residue chain sorting** (`residue_chain=True`): Orders the matrix by residue index 
        after filtering.

        The method can also return just the interaction count per row/column (`count=True`). If 
        both `thr_interactions` and `selected_items` are provided, an error is raised.

        Args:
            interaction_data (InteractionData): The interaction data to be sorted.
            axis (str, optional): Specifies whether to sort rows ('rows') or columns ('columns'). Defaults to 'rows'.
            thr_interactions (int, optional): Minimum number of interactions required to retain a row/column.
            thr_activity (float, optional): Minimum activity value required to retain a row/column.
            selected_items (int, optional): Number of top rows/columns to keep based on interaction count.
            count (bool, optional): If True, returns the count of interactions instead of modifying the matrix.
            residue_chain (bool, optional): If True, sorts the resulting matrix based on residue order in the chain.
            save (str, optional): File path to save the resulting matrix. Defaults to None.

        Returns:
            InteractionData: The sorted interaction matrix.

        Raises:
            ValueError: If multiple selection criteria (`thr_interactions`, `thr_activity`, `selected_items`) are used simultaneously.
            InvalidAxisException: If an invalid axis is provided.
            ValueError: If the matrix dimensions are insufficient for sorting.
        """
        
        def get_interactions(
                cell: str
                ) -> int:
            """
            Counts the number of interactions in a cell formatted with interaction data.

            Args:
                cell (str): Cell containing interaction data, formatted with '|' separating values.

            Returns:
                int: Total number of interactions in the cell.
            """
            interactions = 0
            sections = cell.split(GROUP_DELIM)
            for index in range(1, len(sections), 2):
                interactions += len(sections[index].split(SAME_DELIM))
            return interactions

        def sort_by_residue(
                matrix: list[list[str]]
                ) -> list[list[str]]:
            """
            Sorts the matrix based on residue indices in the first column.

            Args:
                matrix (list[list[str]]): The matrix to be sorted.

            Returns:
                list[list[str]]: The matrix sorted by residue indices.
            """
            # Validate matrix dimensions
            self._verify_dimensions(matrix=matrix)

            # Determine whether to sort by rows or columns
            axis = self._get_residues_axis(matrix=matrix)
            
            # If sorting by columns, transpose the matrix first
            if axis == 'columns':
                matrix = self.transpose_matrix(matrix)

            # Separate the header from the data rows
            header = matrix[0]
            data_rows = matrix[1:]

            # Sort the data rows based on residue indices
            sorted_data_rows = sorted(data_rows, key=lambda row: int(row[0].replace(" ", "")[3:].split("-")[0]))

            # Combine the header with the sorted data rows
            sorted_matrix = [header] + sorted_data_rows

            # If sorting was by columns, transpose the sorted matrix back
            if axis == 'columns':
                sorted_matrix = self.transpose_matrix(sorted_matrix)

            return sorted_matrix

        self._check_variable_types(
            variables=[interaction_data, axis, thr_interactions, thr_activity, selected_items, count, residue_chain, save], 
            expected_types=[InteractionData, str, (int, None.__class__), (float, None.__class__), (int, None.__class__), bool, bool, (str, None.__class__)], 
            variable_names=['interaction_data', 'axis', 'thr_interactions', 'thr_activity', 'selected_items', 'count', 'residue_chain', 'save']
        )

        data = copy.deepcopy(interaction_data)
        matrix = data.matrix

        self._verify_dimensions(matrix=matrix)

        if axis not in ['rows', 'columns']:
            raise InvalidAxisException(axis)
    
        # Raise an error if `thr_interactions` and `selected_items` are provided simultaneously
        if thr_interactions is not None and selected_items is not None:
            raise ValueError("Cannot select by both 'thr_interactions' and 'selected_items' at the same time.")
        if thr_interactions is not None and thr_activity is not None:
            raise ValueError("Cannot select by both 'thr_interactions' and 'thr_activity' at the same time.")
        if thr_activity is not None and selected_items is not None:
            raise ValueError("Cannot select by both 'thr_activity' and 'selected_items' at the same time.")
        
        # Ensure the correct axis is selected for activity-based selection
        if thr_activity is not None:
            axis = 'rows' if self._get_residues_axis(matrix=matrix) == 'columns' else 'columns'

        # Transpose the matrix if operating on columns
        if axis == 'columns':
            matrix = self.transpose_matrix(matrix)
        
        # Initialize a dictionary to store interaction counts per row/column
        reactives = {}
        for row in range(1, len(matrix)):
            for column in range(1, len(matrix[row])):
                cell = matrix[row][column]
                interactions = get_interactions(cell)
                reactives[row] = reactives.get(row, 0) + interactions
        
        # If `count` is True, return the interaction counts
        if count:
            data = [list(reactives.keys()), list(reactives.values())]
            for index in reactives.keys():
                original_value = matrix[index][0]
                split_value = original_value.split('_')[0].strip()  # Splits and removes spaces
                data[0][index-1] = split_value
            return data
        
        # Select rows/columns based on the provided criteria
        elif thr_interactions is not None:
            reactives = [key for key, value in sorted(reactives.items(), key=lambda item: item[1], reverse=True) if value >= thr_interactions]
        elif self._get_residues_axis(matrix) == "columns" and thr_activity is not None:
            reactives = [key for key, value in sorted(reactives.items(), key=lambda item: float(matrix[item[0]][0].split(" (")[1].replace(")", "")), reverse=True) if float(matrix[key][0].split(" (")[1].replace(")", "")) >= thr_activity]
        elif selected_items:
            selected_items = min(selected_items, len(matrix))
            reactives = [key for key, value in sorted(reactives.items(), key=lambda item: item[1], reverse=True)[:selected_items]]
        else:
            reactives = [key for key, value in sorted(reactives.items(), key=lambda item: item[1], reverse=True)]
        
        # Create the selection matrix with the chosen rows/columns
        selection = [matrix[0]] + [matrix[row] for row in reactives]

        # Sort the selection by residue chain if specified
        if residue_chain:
            selection = sort_by_residue(matrix=matrix)

        # Transpose the selection back if it was initially transposed for columns
        if axis == 'columns':
            selection = self.transpose_matrix(selection)

        data.matrix = selection

        if save:
            self.save_interaction_data(interaction_data=data, filename=save)

        return data
    
    def transpose_matrix(
            self, 
            interaction_data: InteractionData, 
            save: str = None
            ) -> InteractionData:
        """
        Transposes the given interaction matrix.

        This method swaps rows and columns in the interaction matrix, effectively 
        transposing it. If a save path is provided, the transposed matrix is stored 
        as a file.

        Args:
            interaction_data (InteractionData): The interaction data containing the matrix to be transposed.
            save (str, optional): File path to save the transposed matrix. Defaults to None.

        Returns:
            InteractionData: The updated InteractionData object with the transposed matrix.

        Raises:
            TypeMismatchException: If the provided arguments have incorrect types.
            ValueError: If the matrix dimensions are invalid.
        """
        
        self._check_variable_types(
            variables=[interaction_data, save], 
            expected_types=[InteractionData, (str, None.__class__)], 
            variable_names=['interaction_data', 'save']
        )

        data = copy.deepcopy(interaction_data)
        matrix = data.matrix

        self._verify_dimensions(matrix=matrix)

        # Transpose the matrix using list comprehension
        transposed = [[row[i] for row in matrix] for i in range(len(matrix[0]))]

        data.matrix = transposed

        if save:
            self.save_interaction_data(interaction_data=data, filename=save)

        return data

##############
# Exceptions #
##############

class TypeMismatchException(Exception):
    def __init__(self, variable_name, expected_types, actual_type):
        expected_types_str = ", ".join([t.__name__ for t in expected_types])
        self.message = f"Variable '{variable_name}' has type {actual_type.__name__}, expected one of ({expected_types_str})."
        super().__init__(self.message)

class FileOrDirectoryException(Exception):
    def __init__(self, path, error_type, message=None):
        self.path = path
        self.error_type = error_type  # Puede ser 'not_found' o 'empty'
        
        default_messages = {
            'not_found': f"File or directory not found: '{path}'",
            'empty': f"Directory '{path}' is empty"
        }
        
        self.message = message if message else default_messages.get(error_type, "An error occurred")
        super().__init__(self.message)

class InvalidColorException(Exception):
    """
    Exception raised when an invalid hexadecimal color is provided.
    """
    def __init__(self, invalid_colors: list[str]):
        self.invalid_colors = invalid_colors
        self.message = f"Invalid hexadecimal color(s) detected: {', '.join(invalid_colors)}"
        super().__init__(self.message)

class InvalidFileExtensionException(Exception):
    """Exception raised when the file extension is not .csv."""
    def __init__(self, filename, message="Invalid file extension. Only '.csv' files are allowed"):
        self.filename = filename
        self.message = f"{message}: '{filename}'"
        super().__init__(self.message)

class InvalidAxisException(Exception):
    """Exception raised for invalid axis values."""
    def __init__(self, axis_value):
        self.axis_value = axis_value
        self.message = f"Invalid axis value: '{axis_value}'. Expected 'rows' or 'columns'."
        super().__init__(self.message)

class InvalidFilenameException(Exception):
    """Exception raised for invalid mode values."""
    def __init__(self, filenames):
        self.filenames = filenames
        output = ""
        for filename in self.filenames:
            output += "\n\t- " + filename
        self.message = f"Some files have an invalid file name: {output}. \nFile names must not contain spaces."
        super().__init__(self.message)

class HeatmapActivityException(Exception):
    """Exception raised for invalid mode values."""
    def __init__(self):
        self.message = "Heatmap modes' max, min and mean require ligand/complex activities."
        super().__init__(self.message)

class InvalidModeException(Exception):
    """Exception raised for invalid mode values."""
    def __init__(self, mode, expected_values):
        self.message = f"Invalid mode value: '{mode}'. Expected:{expected_values}"
        super().__init__(self.message)
